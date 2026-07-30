"""
Scansiona la struttura di cartelle Audicom ed estrae le URL segnalate
per ciascun codice di errore.

Struttura attesa:
  root/
  ├── GEDI Gruppo Editoriale/   ← siti interni
  │   └── {Testata}/{tipo}/file.xlsx
  └── {AltroGruppo}/            ← terzi Manzoni
      └── {Testata}/{tipo}/file.xlsx

Output: lista di dict con tutte le segnalazioni trovate.
"""

import os
import openpyxl

GEDI_GRUPPO = "GEDI Gruppo Editoriale"

# Cartelle da ignorare (contengono solo PDF)
SKIP_FOLDERS = {"dinamico"}

# File Excel da ignorare (struttura diversa, non fa parte del flusso)
SKIP_FILES = {"Apps_Report_GEDI.xlsx"}


def find_segnalazioni(root_path):
    """
    Scansiona root_path e restituisce una lista di dict:
    {
        'url':       str,
        'errore':    str,   # es. "Errore 21"
        'descr':     str,   # es. "Zero page views"
        'gruppo':    str,   # es. "GEDI Gruppo Editoriale"
        'testata':   str,   # es. "La Repubblica"
        'tipo':      str,   # es. "semi_statico_desktop"
        'is_gedi':   bool,
        'xlsx_path': str,
    }
    """
    segnalazioni = []

    for dirpath, dirnames, filenames in os.walk(root_path):
        # Salta le cartelle che non contengono anomalie
        rel = os.path.relpath(dirpath, root_path)
        parts = rel.split(os.sep)

        # Salta cartelle "dinamico" e altre da ignorare
        if any(p in SKIP_FOLDERS for p in parts):
            dirnames.clear()  # non scendere oltre
            continue

        for filename in filenames:
            if not filename.endswith('.xlsx'):
                continue
            if filename in SKIP_FILES:
                continue

            xlsx_path = os.path.join(dirpath, filename)
            gruppo, testata, tipo = _parse_path_parts(parts)
            is_gedi = (gruppo == GEDI_GRUPPO)

            rows = _extract_from_excel(xlsx_path, gruppo, testata, tipo, is_gedi)
            segnalazioni.extend(rows)

    return segnalazioni


def _parse_path_parts(parts):
    """
    Estrae gruppo, testata e tipo dall'elenco di parti del path relativo.
    Struttura: [gruppo, testata, tipo]
    """
    gruppo  = parts[0] if len(parts) > 0 else "?"
    testata = parts[1] if len(parts) > 1 else "?"
    tipo    = parts[2] if len(parts) > 2 else "?"
    return gruppo, testata, tipo


def _extract_from_excel(xlsx_path, gruppo, testata, tipo, is_gedi):
    """
    Legge tutti gli sheet di un Excel e restituisce le segnalazioni trovate.

    Formato Excel Audicom:
      Riga 1: descrizione errore (es. "Zero page views")
      Riga 2: vuota
      Righe 3+: URL, una per riga
    """
    rows = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [WARN] Impossibile leggere {xlsx_path}: {e}")
        return rows

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        descr = None
        urls_started = False

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                descr = str(row[0]).strip() if row[0] else ""
                continue
            if i == 1:
                continue  # riga vuota
            url = row[0]
            if url and str(url).strip():
                url = str(url).strip()
                rows.append({
                    'url':       url,
                    'errore':    sheet_name,
                    'descr':     descr,
                    'gruppo':    gruppo,
                    'testata':   testata,
                    'tipo':      tipo,
                    'is_gedi':   is_gedi,
                    'xlsx_path': xlsx_path,
                })

    wb.close()
    return rows


def read_app_report(segnalazioni_path):
    """
    Legge Apps_Report_GEDI.xlsx dalla root delle segnalazioni Audicom.
    Restituisce lista di dict (una per app) o [] se il file non esiste.
    """
    path = os.path.join(segnalazioni_path, 'Apps_Report_GEDI.xlsx')
    if not os.path.exists(path):
        return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = None
        rows = []
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c).strip() if c else '' for c in row]
                continue
            if not any(row):
                continue
            rows.append(dict(zip(headers, row)))
        wb.close()
        return rows
    except Exception as e:
        print(f"  [WARN] Impossibile leggere Apps_Report_GEDI.xlsx: {e}")
        return []


def get_unique_urls(segnalazioni):
    """
    Restituisce un dict { url: [segnalazione, ...] } raggruppando
    le segnalazioni per URL unica.
    """
    by_url = {}
    for s in segnalazioni:
        url = s['url']
        if url not in by_url:
            by_url[url] = []
        by_url[url].append(s)
    return by_url


def stampa_riepilogo(segnalazioni):
    """Stampa un riepilogo a console dopo il parsing."""
    from collections import Counter
    errori = Counter(s['errore'] for s in segnalazioni)
    testate = Counter(s['testata'] for s in segnalazioni)
    unique_urls = len(get_unique_urls(segnalazioni))

    print(f"\n{'='*50}")
    print(f"Segnalazioni totali : {len(segnalazioni)}")
    print(f"URL uniche          : {unique_urls}")
    print(f"Testate coinvolte   : {len(testate)}")
    print(f"Errori per tipo:")
    for e, n in errori.most_common():
        print(f"  {e}: {n}")
    print(f"{'='*50}\n")
