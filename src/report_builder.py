"""
Costruisce i file Excel di report.

Due file separati:
  - nielsen_gedi_{timestamp}.xlsx    : siti interni GEDI — verifica config TLH
  - nielsen_manzoni_{timestamp}.xlsx : editori terzi Manzoni — verifica SDK + ping
"""

import os
from datetime import datetime
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COL_HEADER_BG = "1F4E79"
COL_HEADER_FG = "FFFFFF"
COL_OK        = "C6EFCE"
COL_KO        = "FFC7CE"
COL_WARN      = "FFEB9C"
COL_GRAY      = "D0D0D0"   # riga da ignorare: http, 404, errore, redirect, articolo vecchio

_THIN_SIDE   = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)


def build_reports(segnalazioni, tlh_results, playwright_results, output_path, tipo=None):
    """
    Genera i file Excel e restituisce (gedi_path, manzoni_path).
    Se tipo='gedi' genera solo il report GEDI (manzoni_path=None).
    Se tipo='manzoni' genera solo il report Manzoni (gedi_path=None).
    """
    os.makedirs(output_path, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    seg_gedi    = [s for s in segnalazioni if s['is_gedi']]
    seg_manzoni = [s for s in segnalazioni if not s['is_gedi']]

    gedi_path    = _build_gedi_report(seg_gedi, tlh_results, playwright_results, output_path, timestamp) if tipo != "manzoni" else None
    manzoni_path = _build_manzoni_report(seg_manzoni, playwright_results, output_path, timestamp)       if tipo != "gedi"    else None

    return gedi_path, manzoni_path


# ─────────────────────────────────────────────
# Helper: righe da evidenziare in grigio
# ─────────────────────────────────────────────

def _should_gray_row(pw):
    """
    Restituisce True per le righe che devono essere grigie:
    Solo per le URL http:// che redirigono su https:// — non sono gestibili lato TLH
    e vanno escluse dalle segnalazioni Audicom. Tutti gli altri casi (errori, timeout,
    HTTP 400+) rimangono visibili normalmente con la nota nella colonna Note.
    """
    return bool(pw.get('http_to_https'))


def _format_pw_error(error_str):
    """Trasforma i messaggi d'errore Playwright in testo leggibile."""
    if not error_str:
        return ""
    if 'Timeout' in error_str or 'timeout' in error_str:
        return "Timeout di navigazione: impossibile verificare la presenza del TLH/SDK/ping"
    if 'ERR_CONNECTION_RESET' in error_str:
        return "Errore di connessione: server non raggiungibile (ERR_CONNECTION_RESET)"
    if 'ERR_NAME_NOT_RESOLVED' in error_str:
        return "Errore di connessione: dominio non risolvibile"
    if 'ERR_CONNECTION_REFUSED' in error_str:
        return "Errore di connessione: connessione rifiutata"
    return f"Errore Playwright: {error_str}"


# ─────────────────────────────────────────────
# File 1: GEDI — verifica config TLH
# ─────────────────────────────────────────────

def _build_gedi_report(segnalazioni, tlh_results, playwright_results, output_path, timestamp):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for errore_code, items in sorted(_group_by_errore(segnalazioni).items()):
        descr = items[0]['descr'] if items else ""
        ws = wb.create_sheet(title=f"{errore_code} - {descr}"[:31])
        _fill_gedi_sheet(ws, items, tlh_results, playwright_results, errore_code)

    ws_rip = wb.create_sheet(title="Riepilogo", index=0)
    _fill_gedi_riepilogo(ws_rip, segnalazioni, tlh_results, playwright_results)

    filepath = os.path.join(output_path, f"nielsen_gedi_{timestamp}.xlsx")
    wb.save(filepath)
    print(f"[report] GEDI: {filepath}")
    return filepath


def _fill_gedi_sheet(ws, segnalazioni, tlh_results, playwright_results, errore_code=""):
    headers = [
        "URL",
        "Testata",
        "TLH in pagina",
        "Config TLH trovata",
        "Mapping Nielsen",
        "Soluzione",
        "SDK in pagina",
        "Ping inviato",
        "Note",
        "Tipo accesso",
        "Nielsen Static URL",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for col_idx in range(1, len(headers) + 1):
        ws.cell(1, col_idx).border = _THIN_BORDER

    note_col = headers.index("Note") + 1

    for row_idx, (url, items) in enumerate(_dedup_by_url(segnalazioni).items(), start=2):
        testate = sorted(set(i['testata'] for i in items))
        tipi    = sorted(set(i['tipo']    for i in items))

        tlh         = tlh_results.get(url, {})
        pw          = playwright_results.get(url, {})
        tlh_matched = tlh.get('matched',     False)
        has_nielsen = tlh.get('has_nielsen', False)
        tlh_in_page = pw.get('tlh_loaded',  False)
        sdk_loaded       = pw.get('sdk_loaded',       False)
        sdk_appid_invalid= pw.get('sdk_appid_invalid', False)
        ping_sent        = pw.get('ping_sent',         False)
        ping_count       = pw.get('ping_count',        0)
        pw_skipped  = bool(pw.get('skipped_reason') or pw.get('http_to_https'))
        pw_error    = bool(pw.get('error'))
        # N/A se URL saltata intenzionalmente O se Playwright non ha potuto verificare (errore/timeout)
        pw_not_verified = pw_skipped or pw_error
        is_errore22 = (errore_code == "Errore 22")

        # Soluzione: solo se la pagina è stata verificata correttamente
        soluzione = ""
        if not pw_not_verified and not tlh.get('error'):
            if is_errore22:
                if ping_count >= 2:
                    soluzione = (
                        f"Doppia inizializzazione Nielsen: {ping_count} ping rilevati in 30s. "
                        f"Verificare che lo snippet Nielsen non venga eseguito due volte (TLH, template, tag manager)"
                    )
                elif ping_count == 0:
                    soluzione = "Errore non riprodotto: nessun ping rilevato nella finestra di osservazione (30s)"
                # ping_count == 1: comportamento corretto, nessuna azione
            else:
                if not tlh_in_page:
                    soluzione = "Inserire TLH in pagina"
                elif not tlh_matched:
                    soluzione = "Aggiungere config TLH"
                elif not has_nielsen:
                    soluzione = "Aggiungere mapping Nielsen"
                elif sdk_loaded and sdk_appid_invalid:
                    soluzione = "AppId Nielsen non definito: aggiungere/correggere regexp nel mapping Nielsen"
                elif not sdk_loaded:
                    soluzione = "Aggiungere regexp nel mapping Nielsen"

        note_parts = []
        if pw.get('service_note'):
            note_parts.append(pw['service_note'])
        if pw.get('skipped_reason'):
            note_parts.append(pw['skipped_reason'])
        if pw.get('http_to_https') and '/corporate' not in url:
            note_parts.append("Redirect HTTP->HTTPS: escludere dalle segnalazioni Audicom")
        if tlh.get('error'):
            note_parts.append(f"TLH err: {tlh['error']}")
        if pw.get('error'):
            note_parts.append(_format_pw_error(pw['error']))
        if sdk_appid_invalid:
            note_parts.append("SDK caricato con appId non definito (conf/undefined.js)")
        if pw.get('final_url') and not pw.get('http_to_https') and '/corporate' not in url:
            note_parts.append(f"Redirect -> {pw['final_url'][:80]}")
        if pw.get('http_status') and pw['http_status'] >= 400:
            note_parts.append(f"HTTP {pw['http_status']}: pagina non raggiungibile")

        row = [
            url,
            ", ".join(testate),
            "N/A" if pw_not_verified else _yesno(tlh_in_page),
            _yesno(tlh_matched),
            _yesno(has_nielsen) if tlh_matched else "",
            soluzione,
            "N/A" if pw_not_verified else _yesno(sdk_loaded),
            "N/A" if pw_not_verified else (str(ping_count) if is_errore22 else _yesno(ping_sent)),
            " | ".join(note_parts),
            ", ".join(tipi),
            tlh.get('nielsen_static') or "",
        ]
        ws.append(row)

        # Solo colonna "Note" va a capo; bordo su tutte le celle dati
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.alignment = Alignment(wrap_text=(col_idx == note_col), vertical="top")
            cell.border = _THIN_BORDER

        _fill_cell = lambda col_name, ok, skipped=False: ws.cell(
            row_idx, headers.index(col_name) + 1
        ).__setattr__('fill', PatternFill("solid", fgColor="FFFFFF" if skipped else (COL_OK if ok else COL_KO)))

        _fill_cell("TLH in pagina",    tlh_in_page, skipped=pw_not_verified)
        _fill_cell("Config TLH trovata", tlh_matched)
        if tlh_matched:
            ws.cell(row_idx, headers.index("Mapping Nielsen") + 1).fill = PatternFill(
                "solid", fgColor=COL_OK if has_nielsen else COL_WARN)
        if sdk_appid_invalid and not pw_not_verified:
            ws.cell(row_idx, headers.index("SDK in pagina") + 1).fill = PatternFill("solid", fgColor=COL_WARN)
        else:
            _fill_cell("SDK in pagina", sdk_loaded, skipped=pw_not_verified)
        if is_errore22 and not pw_not_verified:
            # Errore 22: "Ping inviato" mostra il conteggio; verde=1, giallo=0, rosso≥2
            ping22_color = COL_OK if ping_count == 1 else (COL_WARN if ping_count == 0 else COL_KO)
            ws.cell(row_idx, headers.index("Ping inviato") + 1).fill = PatternFill("solid", fgColor=ping22_color)
        else:
            _fill_cell("Ping inviato", ping_sent, skipped=pw_not_verified)
        if soluzione:
            ws.cell(row_idx, headers.index("Soluzione") + 1).fill = PatternFill(
                "solid", fgColor=COL_WARN)

        # Riga grigia per URL da ignorare: http://, errori HTTP >= 400, errori/timeout Playwright
        if _should_gray_row(pw):
            gray_fill = PatternFill("solid", fgColor=COL_GRAY)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = gray_fill

    _autofit_columns(ws)
    ws.freeze_panes = "A2"


def _rip_header_row(ws, label, col_b="", col_c=""):
    """Riga intestazione sezione nel foglio Riepilogo (3 colonne)."""
    ws.append([label, col_b, col_c])
    r = ws.max_row
    fill = PatternFill("solid", fgColor=COL_HEADER_BG)
    font = Font(bold=True, color=COL_HEADER_FG)
    for col in range(1, 4):
        cell = ws.cell(r, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="left" if col == 1 else "center", vertical="center"
        )


def _fill_gedi_riepilogo(ws, segnalazioni, tlh_results, playwright_results):
    all_urls    = set(s['url'] for s in segnalazioni)
    total       = len(all_urls)
    tlh_in_page = sum(1 for u in all_urls if playwright_results.get(u, {}).get('tlh_loaded'))
    tlh_matched = sum(1 for u in all_urls if tlh_results.get(u, {}).get('matched'))
    tlh_nielsen = sum(1 for u in all_urls if tlh_results.get(u, {}).get('has_nielsen'))
    tlh_err     = sum(1 for u in all_urls if tlh_results.get(u, {}).get('error'))
    sdk_ok      = sum(1 for u in all_urls if playwright_results.get(u, {}).get('sdk_loaded'))
    ping_ok     = sum(1 for u in all_urls if playwright_results.get(u, {}).get('ping_sent'))
    by_testata  = Counter(s['testata'] for s in segnalazioni)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    center = Alignment(horizontal="center", vertical="center")

    # ── Intestazione ──────────────────────────────────────────
    ws.append(["Auto Debug Nielsen — Siti interni GEDI", "", ""])
    ws.merge_cells("A1:C1")
    ws["A1"].font      = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.append(["Data analisi", datetime.now().strftime("%d/%m/%Y %H:%M"), ""])
    ws.append(["", "", ""])

    # ── URL ANALIZZATE ────────────────────────────────────────
    _rip_header_row(ws, "URL ANALIZZATE")
    ws.append(["URL totali (uniche)", total, ""])
    ws.append(["", "", ""])

    # ── ANALISI — tabella Con / Senza ─────────────────────────
    _rip_header_row(ws, "ANALISI IN PAGINA E CONFIG", col_b="Con  ✓", col_c="Senza  ✗")
    metrics = [
        ("TLH in pagina",      tlh_in_page, total - tlh_in_page),
        ("Config TLH trovata", tlh_matched, total - tlh_matched),
        ("Mapping Nielsen",    tlh_nielsen, total - tlh_nielsen),
        ("SDK in pagina",      sdk_ok,      total - sdk_ok),
        ("Ping inviato",       ping_ok,     total - ping_ok),
    ]
    for label, con, senza in metrics:
        r = ws.max_row + 1
        ws.append([label, con, senza])
        ws.cell(r, 2).fill      = PatternFill("solid", fgColor=COL_OK)
        ws.cell(r, 2).font      = Font(bold=True)
        ws.cell(r, 2).alignment = center
        ws.cell(r, 3).fill      = PatternFill("solid", fgColor=COL_KO)
        ws.cell(r, 3).font      = Font(bold=True)
        ws.cell(r, 3).alignment = center

    if tlh_err:
        r = ws.max_row + 1
        ws.append(["Errori TLH check", tlh_err, ""])
        ws.cell(r, 2).fill      = PatternFill("solid", fgColor=COL_WARN)
        ws.cell(r, 2).alignment = center

    ws.append(["", "", ""])

    # ── SEGNALAZIONI PER TESTATA ──────────────────────────────
    _rip_header_row(ws, "SEGNALAZIONI PER TESTATA")
    for testata, count in by_testata.most_common():
        ws.append([testata, count, ""])


# ─────────────────────────────────────────────
# File 2: Manzoni — verifica SDK + ping
# ─────────────────────────────────────────────

def _build_manzoni_report(segnalazioni, playwright_results, output_path, timestamp):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for errore_code, items in sorted(_group_by_errore(segnalazioni).items()):
        descr = items[0]['descr'] if items else ""
        ws = wb.create_sheet(title=f"{errore_code} - {descr}"[:31])
        _fill_manzoni_sheet(ws, items, playwright_results)

    ws_rip = wb.create_sheet(title="Riepilogo", index=0)
    _fill_manzoni_riepilogo(ws_rip, segnalazioni, playwright_results)

    filepath = os.path.join(output_path, f"nielsen_manzoni_{timestamp}.xlsx")
    wb.save(filepath)
    print(f"[report] Manzoni: {filepath}")
    return filepath


def _fill_manzoni_sheet(ws, segnalazioni, playwright_results):
    headers = [
        "URL",
        "Gruppo",
        "Testata",
        "Tipo accesso",
        "SDK in pagina",
        "Ping inviato",
        "Soluzione",
        "Note",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    note_col = headers.index("Note") + 1

    for row_idx, (url, items) in enumerate(_dedup_by_url(segnalazioni).items(), start=2):
        gruppi  = sorted(set(i['gruppo']  for i in items))
        testate = sorted(set(i['testata'] for i in items))
        tipi    = sorted(set(i['tipo']    for i in items))

        pw               = playwright_results.get(url, {})
        sdk_loaded       = pw.get('sdk_loaded',       False)
        sdk_appid_invalid= pw.get('sdk_appid_invalid', False)
        ping_sent        = pw.get('ping_sent',         False)
        pw_skipped       = bool(pw.get('skipped_reason') or pw.get('http_to_https'))
        pw_error         = bool(pw.get('error'))
        pw_not_verified  = pw_skipped or pw_error

        soluzione = ""
        if not pw_not_verified:
            if sdk_loaded and sdk_appid_invalid:
                soluzione = "Correggere appId Nielsen"
            elif not sdk_loaded:
                soluzione = "Integrare SDK Nielsen"
            elif not ping_sent:
                soluzione = "SDK presente ma ping non inviato"

        note_parts = []
        if pw.get('service_note'):
            note_parts.append(pw['service_note'])
        if pw.get('skipped_reason'):
            note_parts.append(pw['skipped_reason'])
        if pw.get('http_to_https') and '/corporate' not in url:
            note_parts.append("URL HTTP con redirect HTTPS: da escludere dalle segnalazioni")
        if pw.get('error'):
            note_parts.append(_format_pw_error(pw['error']))
        if sdk_appid_invalid:
            note_parts.append("SDK caricato con appId non definito (conf/undefined.js)")
        if pw.get('final_url') and not pw.get('http_to_https') and '/corporate' not in url:
            note_parts.append(f"Redirect -> {pw['final_url'][:80]}")
        if pw.get('http_status') and pw['http_status'] >= 400:
            note_parts.append(f"HTTP {pw['http_status']}: pagina non raggiungibile")

        row = [
            url,
            ", ".join(gruppi),
            ", ".join(testate),
            ", ".join(tipi),
            "N/A" if pw_not_verified else _yesno(sdk_loaded),
            "N/A" if pw_not_verified else _yesno(ping_sent),
            soluzione,
            " | ".join(note_parts),
        ]
        ws.append(row)

        # Solo colonna "Note" va a capo; allineamento verticale top su tutte
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(
                wrap_text=(col_idx == note_col), vertical="top"
            )

        sdk_col  = headers.index("SDK in pagina") + 1
        ping_col = headers.index("Ping inviato") + 1
        sol_col  = headers.index("Soluzione") + 1
        if pw_not_verified:
            ws.cell(row_idx, sdk_col).fill  = PatternFill("solid", fgColor="FFFFFF")
            ws.cell(row_idx, ping_col).fill = PatternFill("solid", fgColor="FFFFFF")
        else:
            if sdk_appid_invalid:
                ws.cell(row_idx, sdk_col).fill = PatternFill("solid", fgColor=COL_WARN)
            else:
                ws.cell(row_idx, sdk_col).fill = PatternFill("solid", fgColor=COL_OK if sdk_loaded else COL_KO)
            ws.cell(row_idx, ping_col).fill = PatternFill("solid", fgColor=COL_OK if ping_sent else COL_KO)
        if soluzione:
            ws.cell(row_idx, sol_col).fill = PatternFill("solid", fgColor=COL_WARN)

        # Riga grigia per URL da ignorare: http://, errori HTTP >= 400, errori/timeout Playwright
        if _should_gray_row(pw):
            gray_fill = PatternFill("solid", fgColor=COL_GRAY)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = gray_fill

    _autofit_columns(ws)
    ws.freeze_panes = "A2"


def _fill_manzoni_riepilogo(ws, segnalazioni, playwright_results):
    all_urls   = set(s['url'] for s in segnalazioni)
    sdk_ok     = sum(1 for u in all_urls if playwright_results.get(u, {}).get('sdk_loaded'))
    ping_ok    = sum(1 for u in all_urls if playwright_results.get(u, {}).get('ping_sent'))
    pw_err     = sum(1 for u in all_urls if playwright_results.get(u, {}).get('error'))
    by_gruppo  = Counter(s['gruppo'] for s in segnalazioni)

    _SECTIONS = {"URL ANALIZZATE", "VERIFICA PLAYWRIGHT", "SEGNALAZIONI PER GRUPPO"}
    rows = [
        ["Auto Debug Nielsen - Editori terzi Manzoni", ""],
        ["Data analisi", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["", ""],
        ["URL ANALIZZATE", ""],
        ["URL totali (uniche)", len(all_urls)],
        ["", ""],
        ["VERIFICA PLAYWRIGHT", ""],
        ["SDK Nielsen caricato in pagina",    sdk_ok],
        ["Ping Nielsen inviato",              ping_ok],
        ["SDK non trovato",                   len(all_urls) - sdk_ok],
        ["Ping non trovato",                  len(all_urls) - ping_ok],
        ["Errori Playwright",                 pw_err],
        ["", ""],
        ["SEGNALAZIONI PER GRUPPO", ""],
    ]
    for row in rows:
        ws.append(row)
    for gruppo, count in by_gruppo.most_common():
        ws.append([gruppo, count])

    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    for cell in ws["A"]:
        if cell.value and str(cell.value) in _SECTIONS:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def _group_by_errore(segnalazioni):
    result = {}
    for s in segnalazioni:
        result.setdefault(s['errore'], []).append(s)
    return result


def _dedup_by_url(segnalazioni):
    result = {}
    for s in segnalazioni:
        result.setdefault(s['url'], []).append(s)
    return result


def _style_header_row(ws, row_num, num_cols):
    fill = PatternFill("solid", fgColor=COL_HEADER_BG)
    font = Font(bold=True, color=COL_HEADER_FG)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row_num, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit_columns(ws):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _yesno(value):
    return "Sì" if value else "No"
